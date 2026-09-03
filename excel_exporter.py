import io
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def create_excel(data_list: list) -> io.BytesIO:
    """Creates a styled Excel workbook from channel data."""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Styles
    header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    def apply_header_style(ws, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        
    def auto_adjust_columns(ws):
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def format_numbers(ws, start_row, number_cols):
        for row in range(start_row, ws.max_row + 1):
            for col in number_cols:
                cell = ws.cell(row=row, column=col)
                cell.number_format = '#,##0'

    def apply_alternating_colors(ws, start_row, end_col):
        for row_idx in range(start_row, ws.max_row + 1):
            fill = gray_fill if row_idx % 2 == 0 else white_fill
            for col_idx in range(1, end_col + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # Sheet 1: Channel Overview (Comparison)
    ws_overview = wb.create_sheet(title="Channel Overview")
    
    overview_headers = ["Channel Name", "Handle", "Subscribers", "Total Views", "Total Videos", "Shorts Count", "Long Videos Count", "Created Date", "Country", "Description"]
    ws_overview.append(overview_headers)
    apply_header_style(ws_overview, len(overview_headers))
    
    for row_idx, data in enumerate(data_list, start=2):
        channel = data.get('channel', {})
        summary = data.get('summary', {})
        
        row = [
            channel.get('name', ''),
            channel.get('handle', ''),
            channel.get('subscriber_count', 0),
            channel.get('view_count', 0),
            summary.get('total_videos', 0),
            summary.get('total_shorts', 0),
            summary.get('total_longs', 0),
            channel.get('published_at', ''),
            channel.get('country', ''),
            channel.get('description', '')
        ]
        ws_overview.append(row)
        
    format_numbers(ws_overview, 2, [3, 4, 5, 6, 7])
    apply_alternating_colors(ws_overview, 2, len(overview_headers))
    auto_adjust_columns(ws_overview)

    # Helper for video sheets
    def create_video_sheet(title, all_channel_data, type_filter):
        ws = wb.create_sheet(title=title)
        
        headers = ["#", "Channel", "Title", "Video ID", "Duration", "Published Date", "Views", "Likes", "Comments", "Description", "URL"]
        if type_filter == 'all':
            headers.insert(4, "Type")
            
        ws.append(headers)
        apply_header_style(ws, len(headers))
        
        row_count = 1
        for data in all_channel_data:
            channel_name = data.get('channel', {}).get('name', '')
            
            if type_filter == 'shorts':
                videos = data.get('shorts', [])
            elif type_filter == 'long':
                videos = data.get('long_videos', [])
            else:
                videos = data.get('shorts', []) + data.get('long_videos', [])
                
            for video in videos:
                row_count += 1
                row = [
                    row_count - 1,
                    channel_name,
                    video.get('title', ''),
                    video.get('video_id', ''),
                ]
                
                if type_filter == 'all':
                    row.append("Short" if video.get('duration_seconds', 0) <= 60 else "Long")
                    
                row.extend([
                    video.get('duration_formatted', ''),
                    video.get('published_at', ''),
                    video.get('view_count', 0),
                    video.get('like_count', 0),
                    video.get('comment_count', 0),
                    video.get('description', '')
                ])
                
                ws.append(row)
                
                # Add hyperlink formula
                url_col = len(headers)
                video_id = video.get('video_id', '')
                is_short = video.get('duration_seconds', 0) <= 60
                base_url = "https://youtube.com/shorts/" if is_short else "https://youtube.com/watch?v="
                
                formula = f'=HYPERLINK("{base_url}{video_id}", "Link")'
                cell = ws.cell(row=row_count, column=url_col)
                cell.value = formula
                cell.font = Font(color="0563C1", underline="single")
                
        apply_alternating_colors(ws, 2, len(headers))
        
        # Format numbers
        view_col = headers.index("Views") + 1
        like_col = headers.index("Likes") + 1
        comment_col = headers.index("Comments") + 1
        format_numbers(ws, 2, [view_col, like_col, comment_col])
        
        auto_adjust_columns(ws)

    # Sheet 2: Shorts
    create_video_sheet("Shorts", data_list, type_filter='shorts')
    
    # Sheet 3: Long Videos
    create_video_sheet("Long Videos", data_list, type_filter='long')
    
    # Sheet 4: All Videos
    create_video_sheet("All Videos", data_list, type_filter='all')

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
